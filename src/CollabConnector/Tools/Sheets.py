import re
from collections import OrderedDict
import csv
import pandas


class Sheet:
    data = None
    name = None
    entities = False
    def __init__(self, sheet: str|bytes, name: str = None, detect_entities: bool = True):
        self.name = name
        if isinstance(sheet, bytes):
            sheet = sheet.decode()
        if isinstance(sheet, list|dict):
            self.data = sheet
        else:
            if detect_entities:
                if self.detect_entities(sheet):
                    self.entities = True
                    self.data = self.parse_csv_entities(sheet)
                else:
                    self.data = self.parse_csv(sheet)
        # else:
        #     self.data = self.load(sheet, detect_entities)
        # pass

    def raw(self):
        if isinstance(self.data, dict):
            return self.export_csv_entities(self.data).encode()
        else:
            return self.export_csv(self.data).encode()

    @staticmethod
    def get_column_index(keys, key) -> int:
        if isinstance(keys[0], list):
            keys = keys[0]
        return keys.index(key)

    def get_column(self, sheet:list = None, column: str|int = None) -> list:
        if column is None:
            raise Exception("Missing function parameters")
        if sheet is None:
            sheet = self.data

        if isinstance(column, str):
            column = Sheet.get_column_index(sheet, column)

        return_column = []
        for row in sheet:
            return_column.append(row[column])

        return return_column

    @staticmethod
    def detect_entities(sheet: str) -> bool:
        return True if "Entity:" in sheet[0:10] else False

    @staticmethod
    def parse_xls(workbook_name: str) -> list[OrderedDict]:
        from openpyxl import load_workbook

        workbook = load_workbook(workbook_name)
        sheet = workbook.active

        header = [cell.value for cell in sheet[1]]
        data = []

        for row in sheet.iter_rows(min_row=2):
            row_data = {}
            for i, cell in enumerate(row):
                row_data[header[i]] = cell.value
            data.append(OrderedDict(row_data))

        return data

    @staticmethod
    def parse_csv(sheet: str|bytes) -> list[OrderedDict]:
        sheet = sheet.replace('\ufeff', '')
        if isinstance(sheet, bytes):
            sheet = sheet.decode().split("\n")
        elif isinstance(sheet, str):
            if ".csv"  in sheet.lower() or ".txt" in sheet.lower():
                sheet = list(csv.reader(open(sheet, "r")))
            else:
                sheet = sheet.split("\n")

        header_index = []
        for header in sheet[0]:
            header_index.append(header.replace('\ufeff', '').strip())
        return_list = []
        for row in sheet[1:]:
            if row:
                row_dict = {}
                i = 0
                for cell in row:
                    if i < len(header_index):
                        row_dict[header_index[i]] = cell.replace('\ufeff', '').strip()
                        i += 1
                    else:
                        break

                return_list.append(OrderedDict(row_dict))
        return return_list

    @staticmethod
    def parse_csv_entities(sheet: str|bytes) -> dict[list[OrderedDict]]:
        if isinstance(sheet, bytes):
            sheet = sheet.decode().split("\n")
        elif isinstance(sheet, str):
            if ".csv"  in sheet.lower() or ".txt" in sheet.lower():
                list(csv.reader(open(sheet, "r")))
            else:
                sheet = sheet.split("\n")

        return_dict = {}
        entity = "PLACEHOLDER"
        i = 0
        header_index = []
        while i < len(sheet):
            if "Entity:" in sheet[i]:
                entity = sheet[i].split(":")[1].replace(",","").strip()
                return_dict[entity] = []
                i += 1

                header_index = []
                for header in sheet[i]:
                    header_index.append(header.replace('\ufeff', '').strip())
                i += 1

            while len(sheet) > i and "Entity:" not in sheet[i]:
                if sheet[i] and not re.match(",+", sheet[i]):
                    row_dict = {}
                    c = 0
                    for cell in sheet[i]:
                        if c < len(header_index):
                            row_dict[header_index[c]] = cell.replace('\ufeff', '').strip()
                            c += 1
                        else:
                            break

                    return_dict[entity].append(OrderedDict(row_dict))
                i += 1

        return return_dict

    @staticmethod
    def parse_list(sheet: list) -> list[OrderedDict]:
        header_index = []
        for header in sheet[0]:
            header_index.append(header)

        return_list = []
        for row in sheet[1:]:
            if row:
                row_dict = {}
                i = 0
                for cell in row:
                    if i < len(header_index):
                        row_dict[header_index[i]] = cell
                        i += 1
                    else:
                        break

                return_list.append(OrderedDict(row_dict))
        return return_list

    @staticmethod
    def load(sheet: str|bytes,
             detect_entities: bool = False
             ) -> list[OrderedDict]|dict[list[OrderedDict]]:
        # check for filename as sheet and process
        if len(sheet) < 128 and ".xls" in sheet.lower():
            if detect_entities:
                raise Exception("Splitting Entities is not supported within XLS")
            return Sheet.parse_xls(sheet)
        elif len(sheet) < 128 and (".csv" in sheet.lower() or ".txt" in sheet.lower()) :
            if detect_entities:
                return Sheet.parse_csv_entities(sheet)
            else:
                return Sheet.parse_csv(sheet)

        if isinstance(sheet, bytes):
            sheet = sheet.decode()

        if isinstance(sheet, str):
            if detect_entities and Sheet.detect_entities(sheet):
                return Sheet.parse_csv_entities(sheet)
            else:
                return Sheet.parse_csv(sheet)

        if isinstance(sheet, list):
            return Sheet.parse_list(sheet)

        raise Exception("Cannot load sheet.  Must be CSV, XLS, XLSX or 2-D List")

    @staticmethod
    def export_csv_entities(sheet: dict[list[OrderedDict]] | list[OrderedDict], entities: list|str = None) -> str:
        return_str = ""
        if entities is None:
            entities = list(sheet.keys())
        if isinstance(entities, str):
            entities = [entities]
        for sheet_entity in entities:
            return_str += f"Entity:{sheet_entity}\n"
            return_str += ",".join(sheet[sheet_entity][0].keys())
            return_str += f"\n"
            for row in sheet[sheet_entity]:
                return_str += ",".join(row.values())
                return_str += f"\n"
            return_str += f"\n"
        return return_str

    @staticmethod
    def export_csv(sheet: dict[list[OrderedDict]]|list[OrderedDict], entities: list = None) -> str:
        if entities:
            return Sheet.export_csv_entities(sheet, entities)
        return_str = ",".join(sheet[0].keys())
        return_str += f"\n"
        for row in sheet:
            return_str += ",".join(row.values())
            return_str += f"\n"
        return return_str

    @staticmethod
    def export_xls(sheet: list[OrderedDict], export_file: str) -> bool:
        import openpyxl
        fieldnames = list(sheet[0].keys())

        # create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active

        # append headers
        ws.append(fieldnames)

        # append data
        # iterate `list` of `dict`
        for row in sheet:
            # create a `generator` yield product `value`
            # use the fieldnames in desired order as `key`
            values = (row[k] for k in fieldnames)
            # append the `generator values`
            ws.append(values)

        wb.save(export_file)

        return True

    def export(self,
               sheet: list = None,
               export_file: str = None,
               export_format: str = "CSV",
               ) -> str|bool:
        if sheet is None:
            sheet  = self.data
        if export_format.upper() == "STR" or export_format.upper() == "STRING":
            return Sheet.export_csv(sheet)

        elif export_format.upper() == "CSV":
            if export_file:
                df = pandas.DataFrame(sheet)
                df.to_csv(export_file, index=False)
                # with open(export_file, "w") as write_file:
                #     write_file.write(Sheet.export_csv(sheet))
                return True
            else:
                return Sheet.export_csv(sheet)

        elif export_format.upper() == "XLSX" or export_format.upper() == "XLS":
            if export_file is None:
                raise Exception("Must provide an output filename for XLSX export")
            return Sheet.export_xls(sheet, export_file)

        return False

    @staticmethod
    def compare(value_a: str|int|float|bool,
                value_b: str|int|float|bool|list,
                operator: str = "==",
                case_insensitive: bool = False
                ) -> bool:

        if operator.lower() == "in" and ( isinstance(value_b, list) or isinstance(value_b, str) ):
            return True if value_a in value_b else False

        if operator in ["=", "=="]:
            if isinstance(value_a, str) and isinstance(value_b, str):
                if case_insensitive:
                    return True if value_a.lower() == value_b.lower() else False
                else:
                    return True if value_a == value_b else False
            elif isinstance(value_a, bool) and isinstance(value_b, bool):
                return True if value_a == value_b else False
            else:
                return True if float(value_a) == float(value_b) else False

        if operator == "<":
            return True if float(value_a) < float(value_b) else False
        if operator == "<=":
            return True if float(value_a) <= float(value_b) else False
        if operator == ">":
            return True if float(value_a) > float(value_b) else False
        if operator == ">=":
            return True if float(value_a) >= float(value_b) else False
        if operator == "!=" or operator == "<>":
            return True if float(value_a) >= float(value_b) else False

        raise Exception(f"Comparison operator `{operator}` or values `{value_a}`, `{value_b}` not valid,",
                        "use: ==, <=, >=, <, >, != and `in`")

    @staticmethod
    def filter(sheet: list[OrderedDict] | list[list],
               column: str|int = None,
               value: list|str = None,
               expr: str = None
               ) -> list[OrderedDict]|list[list]:
        if not ( expr or value):
            raise Exception("Must supply either `column` and `filter` value or `expr`")

        return_list = []
        comparison = "=="
        if expr:
            column, value = re.split(r'[<>=]{1,2}', expr)
            comparison = re.sub(r'[^<>=]{1,2}', '', expr)

        for row in sheet:
            if Sheet.compare(row[column], value, comparison):
                return_list.append(row)

        return return_list

    def columns(self, sheet: dict[list[OrderedDict]]|list[OrderedDict] = None, entity: str = None) -> list:
        if sheet is None:
            sheet = self.data
        if isinstance(sheet, dict):
            if entity is None:
                raise Exception("Must Supply Entity for entity type sheets")
            else:
                sheet = sheet[entity]
        return list(sheet[0].keys())

    def column(self,
               sheet: dict[list[OrderedDict]]|list[OrderedDict] = None,
               column: str = None,
               entity: str = None
               ) -> list:
        if column is None:
            raise Exception("Missing function parameters")
        if sheet is None:
            sheet = self.data

        if isinstance(sheet, dict):
            if entity is None:
                raise Exception("Must Supply Entity for entity type sheets")
            else:
                sheet = sheet[entity]

        if column not in self.columns(sheet):
            raise Exception(f"Column not found in sheet: `{column}`")

        return_list = []
        for row in sheet:
            return_list.append(row[column])

        return return_list

    def find_column(self, sheet: list[OrderedDict] = None,
                    column: str = None,
                    partial_match: bool = False,
                    case_insensitive: bool = False,
                    entity: str = None
                    ) -> list:
        if column is None:
            raise Exception("Missing function parameters")
        if sheet is None:
            sheet = self.data
        if isinstance(sheet, dict):
            if entity is None:
                raise Exception("Must Supply Entity for entity type sheets")
            else:
                sheet = sheet[entity]

        return_list = []
        for col in sheet[0].keys():
            if Sheet.compare(column,
                             col,
                             operator = "in" if partial_match else "==",
                             case_insensitive = case_insensitive):
                return_list.append(col)
        return return_list

    def row(self, sheet: list[OrderedDict] = None, row: int = None, entity: str = None) -> OrderedDict:
        if row is None:
            raise Exception("Missing function parameters")
        if sheet is None:
            sheet = self.data
        if isinstance(sheet, dict):
            if entity is None:
                raise Exception("Must Supply Entity for entity type sheets")
            else:
                sheet = sheet[entity]

        if len(sheet) < row:
            raise Exception(f"Row index out of range: `{row}`")
        return sheet[row]

    @staticmethod
    def replace(sheet: list[OrderedDict],
                column: str,
                match: str,
                replacement: str,
                operator: str = "==",
                case_insensitive: bool = False
                ) -> list[OrderedDict]:
        for row in sheet:
            if Sheet.compare(row[column], match, operator, case_insensitive):
                row[column] = replacement

        return sheet

    @staticmethod
    def update(sheet: list[OrderedDict],
                column: str,
                match: str,
                replacement: str,
                operator: str = "==",
                case_insensitive: bool = False
                ) -> list[OrderedDict]:
        for row in sheet:
            if Sheet.compare(row[column], match, operator, case_insensitive):
                row[column] = replacement

        return sheet

    @staticmethod
    def set(sheet: list[OrderedDict],
            column: str,
            replacement: str
            ) -> list[OrderedDict]:
        for row in sheet:
            row[column] = replacement

        return sheet

    @staticmethod
    def substitute(sheet: list[OrderedDict],
                column: str,
                match: str,
                replacement: str
                ) -> list[OrderedDict]:
        for row in sheet:
            if Sheet.compare(row[column], match, replacement):
                row[column] = re.sub(match, replacement, row['column'])

        return sheet