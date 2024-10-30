import sys
import threading
import time
import getpass
import re
import io
import openpyxl
import copy

try:
    import netmiko
except:
    print("To Interact with devices via SSH install netmiko.", file=sys.stderr)
    print(f"<< pip3 install netmiko >>", file=sys.stderr)

try:
    import netconf
except:
    print("For some of the config parsing we use NETCONF which isn't installed.", file=sys.stderr)
    print(f"<< pip3 install netconf >>", file=sys.stderr)


class Connect:
    device_login = None
    keep_system_alive = True
    config = None
    interface_status = []
    speed = {'half': [], 'ten': []}
    hostname = None
    domain = None
    ip = {}
    inventory = None
    serial_number = None
    version = None
    platform = None
    dialpeers = {}
    cdp = {}

    def __init__(self, *args, **kwargs):
        if args or kwargs:
            if args and isinstance(args[0], dict):
                self.device_login = args[0]
            elif kwargs:
                self.device_login = kwargs
        if len(args) == 3:
            self.device_login = {"ip": args[0], "username": args[1], "password": args[2]}
        else:
            self.device_login = {'ip': input("IP Address: "), 'username': input("Username: "),
                                 'password': getpass.getpass()}

        self.device_login['device_type'] = 'cisco_ios'
        self.device_login['keepalive'] = 10

        # setup ssh connection to the switch
        try:
            self.netmiko = netmiko.ConnectHandler(**self.device_login)
            self.netmiko.enable()
            self.netmiko.disable_paging(command='terminal length 0', delay_factor=1, cmd_verify=True, pattern=None)

        except Exception as err:
            raise Exception(f"Cannot connect to this device: {err}")

        else:
            self.status = True
            print(f"Connected to device: {self.device_login['ip']}", file=sys.stderr)

            # send find_prompt() in background every 30sec to keep session alive
            keepalive = threading.Thread(target=self.session_keepalive, args=())  # create thread process
            keepalive.daemon = True  # enable run in background
            keepalive.start()  # start keepalives

        self.parse_config()
        self.config_sections = [""]
        i = 0
        while i < len(self.config):
            if self.config[i] != "!":
                self.config_sections[-1] += f"{self.config[i]}\n"
            i += 1
            if i < len(self.config) and re.match("^[^\s].*", self.config[i]):
                self.config_sections.append("")
        self.parse_version()
        self.parse_dialpeers()
        self.parse_serialnumber()

    def exec(self, *commands, auto_confirm=False):
        send_commands = []
        for arg in commands:
            if isinstance(arg, list):
                for each in arg:
                    send_commands.append(each)
            else:
                send_commands.append(arg)

        if self.status is False:
            try:
                self.netmiko.establish_connection()
                self.netmiko.enable()
                self.netmiko.disable_paging(command='terminal length 0', delay_factor=1, cmd_verify=True, pattern=None)
            except Exception as err:
                print(f"Cannot reestablish SSH connection to: {self.device_login['ip']}", file=sys.stderr)
            else:
                self.status = True

        if self.status:
            output = []
            for command in send_commands:
                # if show command use send_command using auto find promptr return
                if re.search("^sh", command):
                    try:
                        command_out = self.netmiko.send_command(command, cmd_verify=True,
                                                                strip_prompt=False).splitlines()  # .splitlines() to plit lines into list

                    except Exception as err:
                        print(f"Error sending commands: {err}", file=sys.stderr)
                        command_out = False

                    if command_out and not re.search("\% Invalid input detected at ", str(command_out)):
                        command_out.pop()
                        output.append('\n'.join(command_out))
                    else:
                        print(f"Syntax Error: {command} - {command_out}", file=sys.stderr)
                        output.append(False)

                else:
                    # else send via command timing to catch confirmations
                    try:
                        command_out = self.netmiko.send_command_timing(command, cmd_verify=True,
                                                                       strip_prompt=False)  # .splitlines() to plit lines into list
                        if auto_confirm:
                            while re.search("\]$", command_out) or re.search("\?$", command_out.strip()):
                                print(f"{command} - {command_out} - Auto Confirming")
                                try:
                                    command_out = self.netmiko.send_command_timing("")
                                except:
                                    print(f"Error sending commands: {err}", file=sys.stderr)
                                    break

                        output.append(command_out)

                    except Exception as err:
                        print(f"Error sending commands: {err}", file=sys.stderr)
                        output.append(False)
            if len(output) == 1:
                return output[0]
            else:
                return output

        else:
            return False

    def running_config(self):
        return self.exec("show running-config")

    def startup_config(self):
        return self.exec("show startup-config")

    def tech_support(self):
        return self.exec("show tech-support")

    def save(self):
        current_time = int(time.time())
        if self.exec(f'copy start {current_time}_Running-Config.txt'):
            if self.exec('copy run start'):
                return True

        print("Error saving IOS config.", file=sys.stderr)
        return False

    def wr(self):
        return self.save()

    def config_t(self, *commands):
        send_commands = []
        for arg in commands:
            if isinstance(arg, list):
                for each in arg:
                    send_commands.append(each)
            else:
                send_commands.append(arg)
        if self.status is False:
            try:
                self.netmiko.establish_connection()
                self.netmiko.enable()
                self.netmiko.disable_paging(command='terminal length 0', delay_factor=1, cmd_verify=True, pattern=None)

            except Exception as err:
                print(f"Cannot reestablish SSH connection to: {self.device_login['ip']}", file=sys.stderr)

            else:
                self.status = True

        if self.status:
            try:
                output = self.netmiko.send_config_set(send_commands)

            except Exception as err:
                print(f"Error sending commands: {err}", file=sys.stderr)
                return False

            else:
                return output
        else:
            return False

    def exit(self):
        self.keep_system_alive = False
        self.netmiko.cleanup(command='exit')
        return True

    def session_keepalive(self):
        while True and self.keep_system_alive:
            time.sleep(30)
            # get prompt place to keep session alive
            try:
                self.netmiko.find_prompt()

            except Exception as err:
                print(f"Lost connection to IOS device: {self.device_login['ip']}", file=sys.stderr)
                break

        self.status = False

    def parse_config(self, config=None):
        if config:
            config = self.fix_file_format(config)
        else:
            if self.config:
                config = self.config
            else:
                self.config = self.fix_file_format(self.running_config())
                config = self.config
        x = 0
        while x < len(config):
            if "ip domain name" in config[x]:
                self.domain = config[x].split(" ")[-1].strip()

            if "hostname" in config[x]:
                self.hostname = config[x].split(" ")[-1].strip()

            if config[x].find("interface") == 0:
                this_interface = "".join(config[x].split(' ')[1:]).replace('\n', '')
                while x < len(config) and "!" not in config[x]:
                    if "ip address" in config[x] and len(config[x].split(" ")) == 5:
                        self.ip[this_interface] = [config[x].split(" ")[-2], config[x].split(" ")[-1].strip()]
                        break
                    x += 1
            x += 1

    def parse_serialnumber(self, config=None):
        if config:
            config = self.fix_file_format(config)
        else:
            if self.inventory is None:
                self.parse_inventory()
            config = self.inventory
        for pid in config:
            if pid['NAME'] == "Chassis":
                self.serial_number = pid['SN']
                self.platform = pid['PID']
                break
        return self.serial_number

    def parse_inventory(self, config=None):
        if config:
            config = self.fix_file_format(config)
        else:
            if self.inventory:
                return self.inventory
            else:
                config = self.fix_file_format(self.exec("show inventory"))
                self.inventory = []
        # Parse show inventory output
        x = 0
        while x < len(config):
            if 'NAME' in config[x]:
                that_pid = {}
                pid = f"{config[x]}, {config[x + 1]}".replace('\n', '')
                i = 1
                this_pid = pid.split('"')
                while i < len(this_pid):
                    this_pid[i] = this_pid[i].replace(",", "")
                    i += 2

                this_pid = '"'.join(this_pid).replace('"', '').split(',')
                for pid_line in this_pid:

                    try:
                        that_pid[pid_line.split(':')[0].strip()] = pid_line.split(':')[1].replace('"',
                                                                                                  '').replace(
                            "\'", '').strip()
                    except Exception as err:
                        print(f"Error: {config[x]} - {pid_line} - {err}")
                        that_pid = None

                self.inventory.append(that_pid)
            x += 1
        return self.inventory

    def parse_version(self, config=None):
        # Find version for show version
        if config:
            config = self.fix_file_format(config)
        else:
            config = self.exec("show version").split("\n")
        x = -1
        while x < len(config):
            x += 1
            if config[x].find("Version ") > -1:
                self.version = config[x].split("Version ")[1].split(" ")[0]
                break
        return self.version

    def parse_cdp(self, config=None):
        # Parse show cdp neighbor detail
        if config:
            config = self.fix_file_format(config)
        else:
            config = self.exec("show cdp neighbor detail").split("\n")
        x = 0
        cdp_neighbor = {}
        while x < len(config[x]):
            if config[x].find("---") == 0:
                if "Interface" in cdp_neighbor.keys():
                    if "Capabilities" in cdp_neighbor.keys():
                        cdp_neighbor['Capabilities'] = cdp_neighbor['Capabilities'].split(' ')
                    self.cdp[cdp_neighbor['Interface']] = cdp_neighbor
                cdp_neighbor = {}

                x += 1

            else:
                if config[x].find("Entry address") == 0:
                    cdp_neighbor['Entry address'] = []
                    x += 1
                    while x < len(config) and re.match(".*IP* address*", config[x]):
                        cdp_neighbor['Entry address'].append(
                            ":".join(config[x].split(':')[1:]).strip().replace('\n', ''))
                        x += 1

                if config[x].find("Management address") == 0:
                    cdp_neighbor['Management address'] = []
                    x += 1
                    while x < len(config) and re.match(".*IP* address*", config[x]):
                        cdp_neighbor['Management address'].append(
                            config[x].split(':')[1].strip().replace('\n', ''))
                        x += 1

                if config[x].find("Version") == 0:
                    x += 1
                    cdp_neighbor['Version'] = config[x].replace('\n', '')
                    while x < len(config) and not config[x].strip() == "":
                        x += 1

                for line in config[x].split(','):
                    if not config[x].strip() == "":
                        cdp_neighbor[line.split(":")[0].strip()] = line.split(":")[1].strip().replace('\n', '')
                x += 1
        return self.cdp

    # # parse show interface status to get half and 10MB links
    # def interface_status(self, show=['ten', 'half']):
    # 	interface_status = self.exec("show interface status")
    # 	i = 0
    # 	while i < len(interface_status):
    # 		if "#" not in interface_status[i] and '/' in interface_status[i]:
    # 			line = re.sub(" +", " ", interface_status[i]).split(' ')
    # 			self.interface_status.append([line[0], line[-3], line[-2]])
    #
    # 			if "half" in line[-3]:
    # 				self.speed['half'].append(line[0])
    #
    # 			if line[-2].split('-')[-1] == '10':
    # 				interface = line[0].replace("Fa", "FastEthernet").replace("Gi", "GigabitEthernet")
    # 				if len(self.cdp) > 0 and interface in self.cdp.keys():
    # 					self.speed['ten'].append(
    # 						f"{line[0]} - {self.cdp[interface]['Platform']} {self.cdp[interface]['Device ID']}")
    # 				else:
    # 					self.speed['ten'].append(line[0])
    #
    # 		elif "#" in interface_status[i]:
    # 			break
    #
    # 		i += 1
    # 	if 'ten' in show:
    # 		print(f"{self.hostname} - Half Duplex Interfaces {self.speed['half']}")
    # 	if 'half' in show:
    # 		print(f"{self.hostname} - 10MB Interfaces {self.speed['ten']}")

    def fix_file_format(self, file):
        if file is None:
            return self.config
        elif type(file) == str:
            return file.split("\n")
        elif type(file) == io.TextIOWrapper:
            return file.readlines()
        elif type(file) == list:
            return file
        else:
            print("Error: IOS.section must be given a text file via string, file or list", file=sys.stderr)
            return []

    def section(self, search, return_type: type = list):
        output = []
        for section in self.config_sections:
            if search in section:
                output.append(section.strip())
        if return_type == str:
            return "\n".join(output)
        else:
            return output

    def include(self, search, file=None):
        if file is None:
            file = self.config
        else:
            file = self.fix_file_format(file)
        return_string = ""
        for line in file:
            if line.lower().find(search.lower()) > -1:
                return_string += line
                return_string += "\n"
        return return_string

    def parse_dialpeers(self, config=None):
        if config:
            dialpeer_config = self.fix_file_format(config)
        else:
            if len(self.dialpeers) == 0:
                dialpeer_config = copy.deepcopy(self.section("dial-peer voice"))

        i = 0
        for dialpeer in dialpeer_config:
            config = dialpeer.split('\n')
            i = 0
            while i < len(config):
                dialpeer = {'tag': config[i].strip().split(" ")[-2],
                            'type': "voip" if "voip" in config[i] else "pots",
                            'status': 'active',
                            'description': '',
                            'destination': '', 'pattern-map dest': "", 'incoming': '', 'target': '',
                            'server-group dest': "", 'source': '', 'preference': "0", 'translation_incoming': "",
                            'translation_outgoing': "", 'call-block': "", 'codec': "codec g711ulaw", 'dtmf-relay': "",
                            'protocol': "H.323", 'huntstop': "disabled", 'vad': "enabled", 'fax-protocol': "",
                            'fax-rate': "", "sip_profile": "", "config": dialpeer}

                i += 1
                while i < len(config):
                    if "description" in config[i]:
                        dialpeer['description'] = config[i].replace(" description ", "")
                    elif "pots" in config[i]:
                        dialpeer['protocol'] = "pots"
                    elif "destination" in config[i]:
                        dialpeer['destination'] = " ".join(config[i].split(" ")[2:])
                        if 'e164-pattern-map' in dialpeer['destination']:
                            dialpeer['pattern-map dest'] = self.section(f"voice class {dialpeer['destination']}\n",
                                                                        return_type=str)
                    elif "incoming called-number" in config[i]:
                        dialpeer['incoming'] = " ".join(config[i].split(" ")[3:])
                    elif "incoming uri" in config[i]:
                        dialpeer['incoming'] = " ".join(config[i].split(" ")[3:])
                    elif "answer-address" in config[i]:
                        dialpeer['incoming'] = config[i].strip()
                    elif "session target" in config[i] or 'server-group' in config[i]:
                        dialpeer['target'] = " ".join(config[i].split(" ")[2:]).replace("target ", "")
                        if "server-group" in dialpeer['target']:
                            dialpeer['server-group dest'] = self.section(f"voice class {dialpeer['target']}\n",
                                                                         return_type=str)
                    elif "shutdown" in config[i]:
                        dialpeer['status'] = 'shutdown'
                    elif 'port' in config[i] or 'trunk-group' in config[i]:
                        dialpeer['target'] = " ".join(config[i].split(" ")[2:])
                    elif 'voice-class sip bind control source-interface ' in config[i]:
                        dialpeer['source'] = config[i].replace("voice-class sip bind control source-interface ", "")
                    elif 'preference' in config[i]:
                        dialpeer['preference'] = config[i].replace(" preference ", "")
                    elif 'translation-profile incoming' in config[i]:
                        dialpeer['translation_incoming'] = config[i].replace(" translation-profile incoming ", "")
                    elif 'translation-profile outgoing' in config[i]:
                        dialpeer['translation_outgoing'] = config[i].replace(" translation-profile outgoing ", "")
                    elif 'codec' in config[i]:
                        dialpeer['codec'] = config[i].strip().split(" ")[-1]
                    elif 'dtmf-relay' in config[i]:
                        dialpeer['dtmf-relay'] = config[i].replace(" dtmf-relay ", "").split(" ")
                    elif 'session protocol' in config[i]:
                        dialpeer['protocol'] = config[i].replace(" session protocol ", "")
                    elif 'huntstop' in config[i] and "no" not in config[i]:
                        dialpeer['huntstop'] = "enabled"
                    elif 'call-block translation-profile' in config[i]:
                        dialpeer['call-block'] = config[i].replace(" call-block translation-profile ", "").strip()
                    elif ' vad' in config[i]:
                        dialpeer['vad'] = "disabled"
                    elif ' fax protocol' in config[i]:
                        dialpeer['fax-protocol'] = config[i].replace(" fax protocol ", "")
                    elif ' fax rate' in config[i]:
                        dialpeer['fax-rate'] = config[i].replace(' fax rate ', '')
                    elif ' voice-class sip profile' in config[i]:
                        dialpeer['sip_profile'] = config[i].strip().split(" ")[-1]
                    i += 1

            self.dialpeers[dialpeer['tag']] = dialpeer
            i += 1

    @staticmethod
    def export(data, destination=None, export_type="XLSX"):
        export_data = []
        if isinstance(data, dict):
            for item in data:
                export_data.append(data[item])
        elif isinstance(data, list):
            export_data = data

        export_keys = list(export_data[0].keys())
        if export_type.upper() == "CSV":
            export_file = f'{",".join(export_keys)}\n'
            for line in export_data:
                export_line = ""
                for field in line:
                    export_line += f"{line[field]},"
                export_file += f'{export_line[:-1]}\n'

            if destination:
                with open(destination, "w") as file:
                    file.write(export_file)
            else:
                return export_file
        elif export_type.upper() == "XLSX" and destination:
            wb = openpyxl.Workbook()
            wb.save(destination)
            ws = wb.active

            ws.append(export_keys)
            r = 2
            for line in export_data:
                c = 1
                for cell in line:
                    ws.cell(r, c, str(line[cell]))
                    c += 1
                r += 1
            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = ws['B2']
            wb.save(destination)
        else:
            return export_data
